#Ennery Invoice Check

import tkinter as tk
import openpyxl, xlsxwriter, csv
from tkinter import filedialog

tkroot = tk.Tk()
tkroot.withdraw()
file1 = filedialog.askopenfilename(title='Choose Dissection Invoice')

invoiceFile = open(file1)
invoiceReader = csv.reader(invoiceFile)
invoiceData = list(invoiceReader)

Invoice_Rows = []

r=0
for row in invoiceData:
    Inv = invoiceData[r][0]
    Part =  invoiceData[r][1]
    Qty = invoiceData[r][2]
    Prc = invoiceData[r][3]
    PO = invoiceData[r][4]
    TLV = invoiceData[r][6]
    ID = (Part +'&'+ Qty)
    Match = "=IFERROR(INDEX('PT50'!D:D,MATCH(Invoice!G"+str(r+1)+",'PT50'!E:E,0)),0)"
    Duplicate = invoiceData[r][7]
    row = (Inv, Part, Qty, Prc, PO, TLV, ID, Match, Duplicate)
    Invoice_Rows.append(row)
    r+=1
    
#get file 2 (PT50 report)

tkroot = tk.Tk()
tkroot.withdraw()
file2 = filedialog.askopenfilename(title='Choose PT50 File')
pt50 = openpyxl.load_workbook(file2)
sheetp = pt50.get_sheet_by_name('Sheet1')

PT50_Rows = []

for row in range(1, (sheetp.max_row) + 1):
    if sheetp['A' + str(row)].value == 80050 or 80053 or 80052:
        Partp = (str(sheetp['N' + str(row)].value).replace(" ",""))
        Desc = sheetp['O' + str(row)].value
        Qtyp = sheetp['M' + str(row)].value
        Pricp = sheetp['P' + str(row)].value
        IDp = (Partp + str(Qtyp))
        rowp = (Partp,Desc,Qtyp,Pricp,IDp)
        PT50_Rows.append(rowp)

r=0
c=0
dissection = xlsxwriter.Workbook('//agsufs01/Supply/Python/Dissection.xlsx')
sheet = dissection.add_worksheet('Invoice')
sheet2 = dissection.add_worksheet('PT50')
for row in Invoice_Rows:
    sheet.write_row(r,c, row)
    r +=1

r=0
for row in PT50_Rows:
    sheet2.write_row(r,c, row)
    r +=1

invoiceFile.close()
dissection.close()

