import openpyxl, csv, os, time, glob, shutil, datetime
from forex_python.converter import CurrencyRates
c = CurrencyRates()

print ('Before starting, make sure all csv invoices are saved in Supply Chain > CSV > New Shipment.')

EDR = openpyxl.load_workbook('G:\\Supply Chain\\Data\\SHIPMENTS\\Electronic Register - SGS.xlsx')

sheet = EDR.get_sheet_by_name('Register')

#sheet.max_row

lship = sheet.cell(row=(sheet.max_row), column=2).value

while True:
    try:
        SHIP = int(input("Please enter a shipment number. The last used number was " + str(lship)+".")) 
    except ValueError:
        print('Try again - make sure you have chosen a four digit number, that is bigger than ' + str(lship) +'.')
    else:
        break #TO DO - Make more rules, like correct number chosen etc 
print("Thanks. Creating folders and doing stuff...")

#add number to ER


directory = "G:\\Supply Chain\\Ship Files\\SHIP" + str(SHIP) + "\\"

if not os.path.exists(directory): #creates new ship folder
    os.makedirs(directory)

print("G:\\Supply Chain\\Ship Files\\SHIP" + str(SHIP) + "\\ has been created to store all documents relating to shipment")

time.sleep(4)

print("Now converting CSV files...")

os.chdir("G:\\Supply Chain\\CSV\Shipment Uploads\\New Shipment\\") 
read_files = glob.glob("*.csv")
with open("invoice.csv", "wb") as outfile:
    for f in read_files:
        with open(f, "rb") as infile:
            outfile.write(infile.read())

master_invoice = (directory)+"SHIP"+str(SHIP)+"_master_invoice.csv" 
			
shutil.move("invoice.csv",master_invoice)

time.sleep(2)

#delfiles = glob.glob("*.csv") 
#for d in delfiles:
 #   os.remove(d)

print("Master invoice created. Now creating ConfirmPOLines...")

os.chdir(directory)

file = "SHIP"+str(SHIP)+".csv"

csvRows = []
r = 0
#exampleFile = open(root.filename)
exampleFile = open(file)
exampleReader = csv.reader(exampleFile)
exampleData = list(exampleReader)
invoice = "invoice"
for row in exampleData:
    if exampleData[r][0] == 'INH':
        example = 0
    elif exampleData[r][0] == 'IDH':
        invoice = exampleData[r][5]
    elif exampleData[r][0] == 'IDD' and exampleData[r][1].isnumeric():
        SHIPinv = "SHIP"+str(SHIP) + "/" + invoice 
        part = exampleData[r][1]
        qty = exampleData[r][3]
        try:
            price = round(float(exampleData[r][6]) // float(exampleData[r][3]),2)
        except ZeroDivisionError:
            price = 0
        po = exampleData[r][7]
        date = datetime.datetime.now().strftime("%Y%m%d")
        drow = (SHIPinv,part,qty,price,po,date)
        print(drow)
        csvRows.append(drow)
    r=r+1

os.chdir("G:\\Supply Chain\\CSV\\Shipment Uploads\\Ready to Load\\")

outputfile = open('ConfirmPOLines.csv', 'w', newline='')
outputwriter = csv.writer(outputfile)
for row in csvRows:
    outputwriter.writerow(row)
outputfile.close()

time.sleep(5)

print("ConfirmPOLines created in Supply Chain > CSV > Shipment Uploads > Ready to Load")

time.sleep(3)

os.chdir(directory)

while True:
    try:
        DMT = str(input('Is this shipment Airfreight or Seafreight? (Enter A or S)'))
    except:
        print("Not a valid option. Try again")
    if DMT == 'A':
        print("Airfreight confirmed. Creating CSV invoice for DHL.")
        time.sleep(5)
        csvRows = []
        r = 0
        exampleFile = open(file)
        exampleReader = csv.reader(exampleFile)
        exampleData = list(exampleReader)
        invoice = "invoice"
        for row in exampleData:
            if exampleData[r][0] == 'INH':
                example = 0
            elif exampleData[r][0] == 'IDH':
                invoice = exampleData[r][5]
            elif exampleData[r][0] == 'IDD' and exampleData[r][1].isnumeric():
                Acc = 'S0819'
                part = exampleData[r][1]
                desc = exampleData[r][2]
                qty = exampleData[r][3]
                tlv = exampleData[r][6]
                gb = 'GB'
                po = exampleData[r][7]
                gbp = 'GBP'
                drow = (Acc,invoice,part,desc,qty,tlv,gb,po,SHIP,gbp)
                print(drow)
                csvRows.append(drow)
    
            r=r+1

        os.chdir("G:\\Supply Chain\\CSV\\Files for Freight Forwarders\\")

        outputfile = open('SHIP'+str(SHIP)+'_DHL.csv', 'w', newline='') #TODO create new and unused doc... 
        outputwriter = csv.writer(outputfile)
        for row in csvRows:
            outputwriter.writerow(row)
        outputfile.close()
        #root.withdraw()
        print('DHL invoice created in Supply Chain > CSV > Files for Freight Forwarders.')

        
    elif DMT == 'S':
        print('Seafreight confirmed.') #TODO create invoices for JAS
    else:
        print("Not a valid option. Please try again")

print("Now I am going to update the Electronic Register...")

directory = "G:\\Supply Chain\\Ship Files\\SHIP" + str(SHIP) + "\\"

os.chdir(directory)

f = open(file)
         
exampleReader = csv.reader(f)
data = list(exampleReader)
invoices = []
r=0
for row in data:
    if data[r][0] == 'IDH':
        invoices.append(data[r][5])
    r=r+1

FCV = []
r=0
for row in data:
    if data[r][0] == 'IDH':
        FCV.append(data[r][9])
    r=r+1

CI = []
r=0
for row in data:
    if data[r][0] == 'IDH':
        CIs = (float(data[r][9]))-(float(data[r][7]))
        CI.append(CIs)
    r=r+1

WH = []
r=0
for row in data:
    if data[r][1] == '8682':
        WH.append(1)
    elif data[r][1] == '8686':
        WH.append(2)
    elif data[r][1] == '8728':
        WH.append(4)
    elif data[r][1] == '8687':
        WH.append(5)
    elif data[r][1] == '18682':
        WH.append(1)
    elif data[r][1] == '18686':
        WH.append(2)
    elif data[r][1] == '18728':
        WH.append(4)
    elif data[r][1] == '18687':
        WH.append(5)
    elif data[r][1] == '8726':
        WH.append(1)
    elif data[r][1] == '8986':
        WH.append(2)
    elif data[r][1] == '8928':
        WH.append(4)
    elif data[r][1] == '8688':
        WH.append(5)
    elif data[r][1] == '18726':
        WH.append(1)
    elif data[r][1] == '18986':
        WH.append(2)
    elif data[r][1] == '18928':
        WH.append(4)
    elif data[r][1] == '18688':
        WH.append(5)
    r=r+1
        
Supplier = []
r=0
for row in data:
    if data[r][1] == '8682':
        Supplier.append(1)
    elif data[r][1] == '8686':
        Supplier.append(1)
    elif data[r][1] == '8728':
        Supplier.append(1)
    elif data[r][1] == '8687':
        Supplier.append(1)
    elif data[r][1] == '18682':
        Supplier.append(24)
    elif data[r][1] == '18686':
        Supplier.append(24)
    elif data[r][1] == '18728':
        Supplier.append(24)
    elif data[r][1] == '18687':
        Supplier.append(24)
    elif data[r][1] == '8726':
        Supplier.append(732)
    elif data[r][1] == '8986':
        Supplier.append(732)
    elif data[r][1] == '8928':
        Supplier.append(732)
    elif data[r][1] == '8688':
        Supplier.append(732)
    elif data[r][1] == '18726':
        Supplier.append(735)
    elif data[r][1] == '18986':
        Supplier.append(735)
    elif data[r][1] == '18928':
        Supplier.append(735)
    elif data[r][1] == '18688':
        Supplier.append(735)
    r=r+1

GBP = c.get_rate('GBP','AUD')
AUD = []
for i in FCV:
   AUD.append(float(i) * GBP)

XR = []
for i in AUD:
    XR.append(GBP)




